"""THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHOR BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN 
ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION 
WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."""

import streamlit as st

import numpy as np
import pandas as pd

import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

import re
import sys

st.set_page_config(page_title="Instron & DIC",page_icon="⏩")

m = st.markdown("""
<style>
div.stButton > button:first-child {
    font-size:16px;font-weight:bold;height:2em;width:7em;
}
</style>""", unsafe_allow_html=True)

st.image('logo_inegi_big.png')
st.title('Instron and DIC File Converter')

st.write("[Instron & DIC Guidelines](https://inegiuppt-my.sharepoint.com/:w:/g/personal/dcardoso_inegi_up_pt/EXSYXf4q_fRBlltb6uXLQKsBcQT2lPHlp3zaNF1FVn90Lg)")

sample_name = st.text_input('Sample Name and Number')

gauge_length = st.number_input("Gauge Length [mm]",0)

width = st.number_input("Sample Width [mm]",min_value=0.0,max_value=50.0,step=1e-3,format="%.2f")

thickness = st.number_input("Sample Thickness [mm]",min_value=0.0,max_value=10.0,step=1e-3,format="%.2f")

check_force = st.checkbox("My Force is in kN",False)


with st.expander("Change Chord Modulus Calculation Strain Ranges"):

    colunas12,colunas22,colunas32 = st.columns(3)

    with colunas12:
        st.caption("Young\'s Modulus")
        youngs_lower_bond = st.number_input("Young\'s Modulus Lower Bond",min_value=0.0000,max_value=0.5000,value=0.0005,step=1e-4,format="%.4f")
        youngs_upper_bond = st.number_input("Young\'s Modulus Upper Bond",min_value=0.0000,max_value=0.5000,value=0.0025,step=1e-4,format="%.4f")
    with colunas32:
        st.caption("Poisson\'s Ratio")
        poisson_lower_bond = st.number_input("Poisson\'s Ratio Lower Bond",min_value=0.0000,max_value=0.5000,value=0.003,step=1e-4,format="%.4f")
        poisson_upper_bond = st.number_input("Poisson\'s Ratio Upper Bond",min_value=0.0000,max_value=0.5000,value=0.015,step=1e-4,format="%.4f")

#True or False LayOut
st.write("Output Excel File Print Options")

col1, col2, col3 = st.columns(3)

with col1:
    logo_f4y = st.checkbox("FIBRE4YARDS Logo",True)
with col2:
    logo_inegi = st.checkbox("INEGI Logo",True)
with col3:    
    plot_grafico = st.checkbox("Plot Stress / Strain (%)",True)

'----------------------------------------------------------------------------------------------------'

area = width * thickness #mm^2

instron_file_ref=st.file_uploader("Choose INSTRON CSV file  | Force according to selection & Displacement in mm")

dic_file_ref=st.file_uploader("Choose DIC CSV file")

run_button=st.button("Run")

if run_button:

    def arranjar_instron (instron_csv_file):#Function - Input: Ficheiro RAW da Instron / Atenção ao nome das Colunas (!!); Output: DataFrame Organizado Instron

        df1_instron = pd.read_csv(instron_csv_file,sep=",",usecols= [1,2], names=["Displacement","Force"],header=6)

        return df1_instron

    def arranjar_dic (dic_csv_file): #Function - Input: Ficheiro RAW do DIC; Output: DataFrame Organizado DIC
        nomedascolunas=["Exx", "Eyy"] 
        df1_dic = pd.read_csv(dic_csv_file,sep=",",usecols= ["exx [1] - Lagrange","eyy [1] - Lagrange"],header=1)
        df1_dic.columns=nomedascolunas
        df1_dic = df1_dic.dropna(axis=0)

        return df1_dic

    def agrupar_final (file_instron,file_dic): #Agrupar DataFrame Instron com DataFrame DIC
        tba=pd.concat([file_instron, file_dic], axis=1)
        #tba=tba.dropna(axis=0)
        return tba

    instron=arranjar_instron (instron_file_ref)
    dic=arranjar_dic(dic_file_ref)

    tabela_final=agrupar_final(instron,dic)
    
    if check_force:
        tabela_final['Force']=tabela_final['Force']*1000 # Transformar força de kN para N caso a opçao estiver selecionada
            
    tabela_final['Tensile Stress']=tabela_final['Force']/area # Cálculo Tensile Stress
    tabela_final['Deformation']=tabela_final['Displacement']/gauge_length # Cálculo Deformation

    cols=['Displacement','Force','Tensile Stress','Deformation','Exx','Eyy'] #Reordenar colunas

    tabela_final=tabela_final[cols] #Reordenar colunas

    #print(tabela_final)

    from scipy import stats

    #Definir intervalos para Eyy onde serão calculados os declives e consequentemente os valores de Young's Modulus e Poisson's Ratio

    #youngs_lower_bond = 0.0005 #Valor de Acordo com a Norma. Verificar!
    #youngs_upper_bond = 0.0025 #Valor de Acordo com a Norma. Verificar!
    #poisson_lower_bond = 0.003 #Valor de Acordo com a Norma. Verificar!
    #poisson_upper_bond = 0.015 #Valor de Acordo com a Norma. Verificar!
    

    #Cálculo Young's Modulus
    tabela_young_modulus=tabela_final.loc[(tabela_final['Eyy']>= youngs_lower_bond) & (tabela_final['Eyy']< youngs_upper_bond)
                                          &(tabela_final['Tensile Stress']<0.9*tabela_final['Tensile Stress'].max())]
    if len(tabela_young_modulus)!=0:
        young_variable=stats.linregress(tabela_young_modulus['Eyy'],tabela_young_modulus['Tensile Stress'])

    #Cálculo Poisson's Ratio
    tabela_poisson=tabela_final.loc[(tabela_final['Eyy']>= poisson_lower_bond) & (tabela_final['Eyy']< poisson_upper_bond)
                                   &(tabela_final['Tensile Stress']<0.9*tabela_final['Tensile Stress'].max())]
    if len(tabela_poisson)!=0:
        poisson_variable = stats.linregress(tabela_poisson['Eyy'],tabela_poisson['Exx'])

    if young_variable.slope>0:
        print('The Young\'s Modulus E is ' + str(round(young_variable.slope/1000,4)) + ' GPa')
    else:
        print('Error Calculating the Young\'s Modulus E. Value cannot be less than 0. Possible DIC error.')

    print('The Poisson\'s Ratio is ' + str(abs(round(poisson_variable.slope,4))))

    st.dataframe(tabela_final)
    
    st.write('Young\'s Modulus = '+ str(round(young_variable.slope/1000,4))+' GPa')
    st.write('Poisson\'s Ratio = ' + str(abs(round(poisson_variable.slope,4))))
    
    fig = px.scatter(tabela_final, x='Eyy', y='Tensile Stress', marginal_y="box",
           marginal_x="box",template="ggplot2")
    fig.update_layout(
        yaxis = dict(
            tickmode = 'linear',
            tick0 = 0,
            dtick = 500,
            tickformat = '.2f'
        )
    )
    st.plotly_chart(fig, use_container_width=True)

    #Criar Ficheiro Excel
    tabela_final.to_excel('tabela_final.xlsx', sheet_name='Test', index=False,float_format="%.5f",startrow=9, startcol=1)

    #Formatar Ficheiro Excel
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.styles import colors
    from openpyxl.styles.borders import Border, Side
    from openpyxl.cell import Cell
    from openpyxl.writer.excel import save_virtual_workbook

    from openpyxl.chart import (
        LineChart,
        BarChart,
        ScatterChart,
        Reference,
        Series,
    )

    wb = openpyxl.load_workbook('tabela_final.xlsx')

    sheet = wb.active

    sheet.merge_cells('B4:G5')
    sheet.merge_cells('B9:C9')
    sheet.merge_cells('D9:E9')
    sheet.merge_cells('F9:G9')

    fontObj1 = Font(bold=True)
    fontObj2 = Font(size=16, bold=True)

    sheet['B4'] = 'Tensile Testing - Raw Data for ' + str(sample_name)

    sheet['J10'] = 'Max. Displacement (mm)'
    sheet['J10'].font = fontObj1
    sheet['M10'] = tabela_final['Displacement'].max()

    sheet['J11'] = 'Max. Force (N)'
    sheet['J11'].font = fontObj1
    sheet['M11'] = tabela_final['Force'].max()

    sheet['J12'] = 'Tensile stress at Maximum Force (MPa)'
    sheet['J12'].font = fontObj1
    sheet['M12'] = tabela_final['Force'].max()/area

    sheet['B9'] = 'INSTRON Data'
    sheet['D9'] = 'DIC Data'
    sheet['F9'] = 'Calculation'

    sheet['B7'] = 'Young Modulus'
    sheet['C7'] = round(young_variable.slope/1000,4)
    sheet['D7'] = 'GPa'

    sheet['F7'] = 'Poissons Ratio'
    sheet['G7'] = abs(round(poisson_variable.slope,4))

    blueFill=PatternFill(fgColor="A5EAEF", fill_type = "solid")
    whiteFill=PatternFill(fgColor="ffffff", fill_type = "solid")

    for row in sheet['A1:Y{}'.format(sheet.max_row)]:
        for cell in row:
            cell.fill = whiteFill
            cell.alignment = Alignment(horizontal='center',vertical='center')


    sheet['B4'].fill = blueFill
    sheet['B4'].font = fontObj2

    if logo_f4y:
        img_f4y = openpyxl.drawing.image.Image('logo_f4y.png')
        img_f4y.anchor = 'B2'
        sheet.add_image(img_f4y)

    if logo_inegi:
        img_inegi = openpyxl.drawing.image.Image('logo_inegi.png')
        img_inegi.anchor = 'G2'
        sheet.add_image(img_inegi)

    sheet.row_dimensions[2].height = 20
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 16
    sheet.column_dimensions['E'].width = 16
    sheet.column_dimensions['F'].width = 16
    sheet.column_dimensions['G'].width = 16

    if plot_grafico:
        c1 = ScatterChart()
        c1.title = 'Tensile Stress vs Strain'

        xvalues = Reference(sheet, min_col=7, min_row=11, max_row=sheet.max_row)
        values = Reference(sheet, min_col=4, min_row=11, max_row=sheet.max_row)
        series = Series(values, xvalues, title_from_data=False)
        c1.series.append(series)

        s1 = c1.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "04939e" # Alterar cor
        s1.marker.graphicalProperties.line.solidFill = "04939e" # Alterar cor
        s1.graphicalProperties.line.noFill = True  # Esconder Linhas
        c1.style = 1
        c1.x_axis.title = 'Strain'
        c1.y_axis.title = 'Tensile Stress (MPa)'
        c1.y_axis.majorGridlines = None
        c1.x_axis.majorGridlines = None
        c1.x_axis.scaling.min = 0
        c1.y_axis.scaling.min = 0
        c1.height = 13
        c1.width = 30

        sheet.add_chart(c1, 'H14')
        
        st.balloons()
    streamtest = save_virtual_workbook(wb)
    st.download_button("Download Final Excel File",streamtest,'Tabela_Final_'+str(sample_name)+'.xlsx')
